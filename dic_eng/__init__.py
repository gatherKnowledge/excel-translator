
from openpyxl import Workbook, load_workbook
import trans as bs

# 작업파일 오픈
wb = load_workbook('C:\\words.xlsx')
ws = wb.active

# soup동작
def getDic(word):
    return bs.trans(word)

def putValue(arr):
    spell = getAlphabet(arr[0][0:1])
    for idx, word in arr :
        ws[spell+idx] = getDic(word)

# 저장 될 장소 찾기
def getAlphabet(spell) :
    alphabet = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N']
    for idx, a in enumerate(alphabet):
        if spell == a :
            if idx+1 <= len(alphabet) :
                return alphabet[idx+1]

def getRowArray(rn):
    rowArray = ws[rn]
    # print(rowArray)
    return rowArray

def saveExcel():
    wb.save('C:\\word_trans.xlsx')

def init() :
    row1 = ws[1]
    arr = list()
    for cell in row1:
        ad = str(cell)
        v = str(cell.value)
        print('Adr : ' + ad + '\tvalue : ' + v)
        if (v != 'None'):
            # print(cell.column)
            arr.append(cell.column)

    for e in arr:
        eachArray = getRowArray(e)
        # 데이터가 있는 첫번째 열
        for aE in eachArray:
            v = str(aE.value)
            if v != 'None':
                print('v' + v)
                translated = getDic(v)
                targetRow = getAlphabet(aE.column)
                target = targetRow + str(aE.row)
                print('result : ' + target)
                ws[target] = translated
            else:
                break
    # 파일 저장
    saveExcel()

if __name__ == '__main__':
    init()
